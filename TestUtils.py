import openpyxl 

result_file = 'Software testing assignment 2.xlsx'
wbk = openpyxl.load_workbook(result_file)

class TestUtil:
    @staticmethod
    def checkTestcase(result,expect,num):
        #lấy sheet paper
        sheet_name = num.split('-')[1]
        sheet_name = TestUtil.get_sheet_name(sheet_name)
        sheet = wbk[sheet_name]
        
        #so sánh kết quả đã ra với expect value
        result = True if result == expect else False

        #lưu vào file excel
        for row in sheet['A']:
            if row.value == num:
                input = "PASSED" if result else "FAILED"
                sheet.cell(row=row.row, column=sheet.max_column, value=input)
                wbk.save('Software testing assignment 2.xlsx')
                break
        return result

    @staticmethod
    def get_sheet_name(i):
        switcher={
            'LI':'Log in',
            'RE':'Log in',
            'PM':'Profile management',
            'JO':'Jobs',
            'CV':'CV templates',
            'VS':'View salary',
            'EC':'Explore company',
            'AM':'Account management'
            }
        return switcher.get(i)
